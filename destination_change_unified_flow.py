"""
Destination Change Unified Flow

Input:
  1) PlanDetailTimeline raw CSV (chua bo 6 dong dau)
  2) Production Schedule raw CSV (chua bo 6 dong dau)
  3) DueDateCalc Excel
  4) Target Week / Wk3

Output:
  Final optimized Excel, co them debug sheets de kiem tra logic.

Default behavior:
  - F Wk3 lay tu Production Schedule, chi lay S/F/P = F.
  - PlanDetailTimeline duoc convert ETA -> ETD bang DueDateCalc.
  - Offset mac dinh la legacy compatible de match SI-SS_WANEK 3.py hien tai.
    Neu muon dung thuan DueDateCalc: chay voi offset_mode="due_date".
"""

from __future__ import annotations

import argparse
import math
import os
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox, simpledialog
except Exception:
    tk = None
    ttk = None
    filedialog = None
    messagebox = None
    simpledialog = None

# ============================================================
# Config
# ============================================================

OUTPUT_COLUMNS = [
    "Item",
    "ProdResourceID",
    "Whse",
    "F Wk3",
    "Sum of SI Wk3",
    "Sum of SI-SS Wk3",
    "Average of SS Wk3",
    "Vendor",
]

DTYPE_MAP = {
    "FIRM DEMAND": "FIRM DEMANDS",
    "FIRM DEMANDS": "FIRM DEMANDS",
    "FIRM POS": "FIRM POS",
    "FIRM PO": "FIRM POS",
    "PLANNED POS": "PLANNED POS",
    "PLANNED PO": "PLANNED POS",
    "SHIPPABLE INV": "SHIPPABLE INV",
    "SHIPPABLE INVENTORY": "SHIPPABLE INV",
    "SAFETY STK": "SAFETY STK",
    "SAFETY STOCK": "SAFETY STK",
    "NET FCST": "NET FCST",
    "NET FORECAST": "NET FCST",
}

# Mapping dung theo SI-SS_WANEK 3.py hien tai.
# Dung de dam bao file output match logic hien tai khi DueDateCalc hien tai duoc dung.
LEGACY_WANEK_OFFSETS = {
    "1": 9,
    "5": 5,
    "12": 8,
    "15": 7,
    "151": 7,
    "16": 8,
    "17": 8,
    "18": 8,
    "19": 8,
    "20": 8,
    "201": 8,
    "242": 5,
    "28": 6,
    "3": 7,
    "335": 1,
    "42": 6,
    "49": 5,
    "60": 7,
    "70": 10,
    "ECR": 7,
    "ALL": 10,
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
VENDOR_FALLBACK_COL = "Vendor"


# ============================================================
# Common helpers
# ============================================================

def normalize_item(value) -> str:
    """Clean Item # values like ="01226" -> 1226."""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    m = re.match(r'^=\s*"(.*)"$', text)
    if m:
        text = m.group(1).strip()
    text = text.strip().strip('"').strip()
    # Ashley exports often keep leading zero as Excel formula text. Existing output uses integer-like item.
    if re.fullmatch(r"0*\d+", text):
        return str(int(text))
    return text


def normalize_whse(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    m = re.match(r'^=\s*"(.*)"$', text)
    if m:
        text = m.group(1).strip()
    try:
        f = float(text)
        if f.is_integer():
            return str(int(f))
    except Exception:
        pass
    return text.strip().upper()


def clean_dtype(series: pd.Series) -> pd.Series:
    s = series.fillna("").astype(str).str.strip().str.upper()
    return s.map(lambda x: DTYPE_MAP.get(x, x))


def parse_user_date(text: str) -> date:
    text = str(text).strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return pd.to_datetime(text).date()


def fmt_date(d: date) -> str:
    return f"{d.month}/{d.day}/{d.year}"


def saturday_of_current_week(today: Optional[date] = None) -> date:
    if today is None:
        today = date.today()
    return today + timedelta(days=(5 - today.weekday()) % 7)


def ensure_unique_output_path(path: str) -> str:
    path = os.path.abspath(path)
    if not os.path.exists(path):
        return path
    folder, filename = os.path.split(path)
    stem, ext = os.path.splitext(filename)
    idx = 1
    while True:
        candidate = os.path.join(folder, f"{stem}_{idx}{ext}")
        if not os.path.exists(candidate):
            return candidate
        idx += 1


def read_report_csv(path: str, dtype=str) -> pd.DataFrame:
    """Read Ashley report CSV that has metadata lines before actual header."""
    header_row = None
    with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
        for i, line in enumerate(f):
            if line.lstrip().startswith("Item #"):
                header_row = i
                break
    if header_row is None:
        raise ValueError(f"Khong tim thay dong header 'Item #' trong file: {path}")
    df = pd.read_csv(path, skiprows=header_row, dtype=dtype, low_memory=False, index_col=False)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def extract_report_date_from_csv(path: str) -> Optional[datetime]:
    with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
        for _ in range(10):
            line = f.readline()
            if not line:
                break
            if "Report Date" in line:
                text = line.split(":", 1)[-1].strip()
                try:
                    return pd.to_datetime(text).to_pydatetime()
                except Exception:
                    return None
    return None


def parse_header_to_date(col_name) -> Optional[date]:
    if isinstance(col_name, (datetime, pd.Timestamp)):
        return pd.to_datetime(col_name).date()
    text = str(col_name).strip()
    try:
        return pd.to_datetime(text).date()
    except Exception:
        return None


def build_date_column_map(df: pd.DataFrame) -> Dict[date, str]:
    mapping = {}
    for c in df.columns:
        d = parse_header_to_date(c)
        if d is not None:
            mapping[d] = c
    return mapping


def date_range_saturdays(start_date: date, end_date: date) -> List[date]:
    out = []
    cur = start_date
    while cur <= end_date:
        out.append(cur)
        cur += timedelta(days=7)
    return out


def get_numeric(df: pd.DataFrame, col) -> pd.Series:
    if col not in df.columns:
        return pd.Series(0.0, index=df.index)
    return pd.to_numeric(df[col], errors="coerce").fillna(0.0)


def group_value(df: pd.DataFrame, key_cols, value_col, output_name) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=key_cols + [output_name])
    out = df.groupby(key_cols, dropna=False, as_index=False)[value_col].sum()
    return out.rename(columns={value_col: output_name})


# ============================================================
# Step 1: DueDateCalc -> warehouse offset
# ============================================================

def load_due_date_offsets(due_date_calc_path: str, offset_mode: str = "legacy_compatible") -> Tuple[Dict[str, int], pd.DataFrame]:
    """
    offset_mode:
      - legacy_compatible: use SI-SS_WANEK 3.py offsets when whse exists there, otherwise ceil(Delivery Days/7).
      - due_date: always use ceil(Delivery Days/7).
    """
    raw = pd.read_excel(due_date_calc_path, sheet_name=0, header=None)
    header_idx = None
    for i in range(len(raw)):
        vals = [str(x).strip() for x in raw.iloc[i].tolist()]
        if "Warehouse" in vals and any("Delivery Days" in v for v in vals):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Khong tim thay header Warehouse / Delivery Days trong DueDateCalc.")

    df = pd.read_excel(due_date_calc_path, sheet_name=0, header=header_idx)
    df.columns = [str(c).strip() for c in df.columns]
    delivery_col = next((c for c in df.columns if "Delivery Days" in c), None)
    if delivery_col is None:
        raise ValueError("DueDateCalc thieu cot Delivery Days.")

    rows = []
    offset_map = {}
    for _, r in df.iterrows():
        warehouse_text = str(r.get("Warehouse", "")).strip()
        if not warehouse_text or warehouse_text.lower() == "nan":
            continue
        whse = normalize_whse(warehouse_text.split("-", 1)[0])
        if not whse:
            continue
        days = pd.to_numeric(r.get(delivery_col), errors="coerce")
        if pd.isna(days):
            continue
        due_calc_offset = max(1, int(math.ceil(float(days) / 7.0)))
        legacy_offset = LEGACY_WANEK_OFFSETS.get(whse)
        if offset_mode == "legacy_compatible" and legacy_offset is not None:
            used_offset = int(legacy_offset)
            source = "SI-SS_WANEK legacy"
        else:
            used_offset = due_calc_offset
            source = "DueDateCalc ceil(days/7)"
        offset_map[whse] = used_offset
        rows.append({
            "Whse": whse,
            "Warehouse": warehouse_text,
            "Delivery Days": float(days),
            "DueDateCalc Offset Weeks": due_calc_offset,
            "Legacy Offset Weeks": legacy_offset,
            "Used Offset Weeks": used_offset,
            "Offset Source": source,
            "Offset Delta Used_vs_DueDate": used_offset - due_calc_offset,
        })
    if not offset_map:
        raise ValueError("Khong doc duoc warehouse offset tu DueDateCalc.")
    return offset_map, pd.DataFrame(rows)


# ============================================================
# Step 2: PlanDetailTimeline ETA -> ETD
# ============================================================

def convert_plan_eta_to_etd(plan_csv_path: str, offset_map: Dict[str, int]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    raw = read_report_csv(plan_csv_path, dtype=str)
    required = ["Item #", "Whse", "Data Type"]
    missing = [c for c in required if c not in raw.columns]
    if missing:
        raise ValueError(f"PlanDetailTimeline thieu cot: {missing}")

    raw = raw.copy()
    raw["Item #"] = raw["Item #"].map(normalize_item)
    raw["Whse"] = raw["Whse"].map(normalize_whse)
    raw["Data Type"] = raw["Data Type"].fillna("").astype(str).str.strip()

    # SI-SS_WANEK 3.py assumes first 3 columns are key, last 20 columns are master data.
    original_date_cols = list(raw.columns[3:-20])
    if not original_date_cols:
        raise ValueError("PlanDetailTimeline khong co cot ngay tuan o vi tri expected.")

    original_dates = []
    for c in original_date_cols:
        try:
            original_dates.append(pd.to_datetime(str(c).strip()).date())
        except Exception as e:
            raise ValueError(f"Cot ngay khong doc duoc trong PlanDetailTimeline: {c}") from e

    # Match SI-SS_WANEK 3.py: extend 22 weeks (154 days) backward.
    extended_dates = [d - timedelta(days=154) for d in original_dates]
    all_dates = extended_dates + original_dates
    date_labels = [fmt_date(d) for d in all_dates]

    out_values = pd.DataFrame(0.0, index=raw.index, columns=date_labels)
    numeric_original = raw[original_date_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)

    unknown_whse = set()
    offset_used_by_row = []
    for whse, idx in raw.groupby("Whse", sort=False).groups.items():
        offset = offset_map.get(whse)
        if offset is None:
            # Safe fallback: no shift. Debug will expose this.
            offset = 0
            unknown_whse.add(whse)
        offset_used_by_row.extend([(i, offset) for i in idx])
        shifted_labels = [fmt_date(d - timedelta(days=7 * offset)) for d in original_dates]
        # Some shifted labels may be outside all_dates if offset > 22; ignore those safely.
        for src_col, dst_label in zip(original_date_cols, shifted_labels):
            if dst_label in out_values.columns:
                out_values.loc[idx, dst_label] = numeric_original.loc[idx, src_col].values

    master_cols = list(raw.columns[-20:]) if len(raw.columns) >= 23 else []
    converted = pd.concat([raw[required].reset_index(drop=True), out_values.reset_index(drop=True), raw[master_cols].reset_index(drop=True)], axis=1)

    debug = pd.DataFrame([
        ["Plan rows", len(raw)],
        ["Original first week", fmt_date(min(original_dates))],
        ["Original last week", fmt_date(max(original_dates))],
        ["Converted first ETD week", fmt_date(min(all_dates))],
        ["Converted last ETD week", fmt_date(max(all_dates))],
        ["Unknown Whse count", len(unknown_whse)],
        ["Unknown Whse list", ", ".join(sorted(unknown_whse))],
    ], columns=["Field", "Value"])
    return converted, debug


# ============================================================
# Step 3: Production Schedule -> F Wk3
# ============================================================

def build_production_date_map(columns: Iterable[str], report_date: Optional[datetime], target_week: date) -> Dict[date, str]:
    date_cols = []
    for c in columns:
        text = str(c).strip()
        if re.fullmatch(r"\d{1,2}/\d{1,2}", text):
            date_cols.append(c)
    if not date_cols:
        raise ValueError("Production Schedule khong co cot tuan dang M/D.")

    base_year = report_date.year if report_date is not None else target_week.year
    # If the first schedule month is much later than target month, it may belong to previous year.
    first_month = int(str(date_cols[0]).strip().split("/")[0])
    year = base_year
    if first_month - target_week.month > 6:
        year -= 1

    mapping = {}
    prev_month = None
    for c in date_cols:
        m, d = [int(x) for x in str(c).strip().split("/")]
        if prev_month is not None and m < prev_month:
            year += 1
        prev_month = m
        mapping[date(year, m, d)] = c
    return mapping


def load_fwk3_from_production(production_csv_path: str, target_week: date) -> Tuple[pd.DataFrame, pd.DataFrame]:
    prod = read_report_csv(production_csv_path, dtype=str)
    prod.columns = [str(c).strip() for c in prod.columns]
    required = ["Item #", "Whse", "S/F/P"]
    missing = [c for c in required if c not in prod.columns]
    if missing:
        raise ValueError(f"Production Schedule thieu cot: {missing}")

    report_dt = extract_report_date_from_csv(production_csv_path)
    date_map = build_production_date_map(prod.columns, report_dt, target_week)
    week_col = date_map.get(target_week)
    if week_col is None:
        available = ", ".join(fmt_date(d) for d in sorted(date_map.keys())[:5]) + " ... " + ", ".join(fmt_date(d) for d in sorted(date_map.keys())[-5:])
        raise ValueError(f"Production Schedule khong co cot target week {fmt_date(target_week)}. Available: {available}")

    prod = prod.copy()
    prod["Item"] = prod["Item #"].map(normalize_item)
    prod["Whse"] = prod["Whse"].map(normalize_whse)
    prod["S/F/P"] = prod["S/F/P"].fillna("").astype(str).str.strip().str.upper()
    prod["F Wk3"] = pd.to_numeric(prod[week_col], errors="coerce").fillna(0.0)

    f = prod[prod["S/F/P"] == "F"].copy()
    grouped = f.groupby(["Item", "Whse"], dropna=False, as_index=False)["F Wk3"].sum()

    debug = pd.DataFrame([
        ["Production rows", len(prod)],
        ["Production F rows", len(f)],
        ["Production report date", str(report_dt) if report_dt else ""],
        ["TargetWeek", fmt_date(target_week)],
        ["Production target column", week_col],
        ["F Wk3 total", float(grouped["F Wk3"].sum())],
        ["F Wk3 nonzero item-whse", int((grouped["F Wk3"] != 0).sum())],
    ], columns=["Field", "Value"])
    return grouped, debug




def build_optimizer_input_direct_from_plan(
    plan_csv_path: str,
    offset_map: Dict[str, int],
    f_wk3: pd.DataFrame,
    target_week: date,
    current_week: date,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Fast path: compute optimizer input directly from raw PlanDetailTimeline without materializing 44-week converted file."""
    raw = read_report_csv(plan_csv_path, dtype=str)
    raw.columns = [str(c).strip() for c in raw.columns]
    required_base = ["Item #", "Whse", "Data Type", "Coll. Class", "MakeBuy Code"]
    missing = [c for c in required_base if c not in raw.columns]
    if missing:
        raise ValueError(f"PlanDetailTimeline thieu cot: {missing}")

    raw = raw.copy()
    raw["Item #"] = raw["Item #"].map(normalize_item)
    raw["Whse"] = raw["Whse"].map(normalize_whse)
    raw["Data Type"] = clean_dtype(raw["Data Type"])
    raw["MakeBuy Code"] = raw["MakeBuy Code"].fillna("").astype(str).str.strip().str.upper()
    raw["Coll. Class"] = raw["Coll. Class"].fillna("").astype(str).str.strip()

    date_cols = list(raw.columns[3:-20])
    if not date_cols:
        raise ValueError("PlanDetailTimeline khong co cot ngay tuan o vi tri expected.")
    date_map = {}
    for c in date_cols:
        try:
            date_map[pd.to_datetime(str(c).strip()).date()] = c
        except Exception as e:
            raise ValueError(f"Cot ngay khong doc duoc trong PlanDetailTimeline: {c}") from e
    original_dates = sorted(date_map.keys())
    first_original_week = min(original_dates)
    last_original_week = max(original_dates)
    first_etd_week = first_original_week - timedelta(days=154)

    raw = raw[raw["MakeBuy Code"] == "B"].copy()
    if raw.empty:
        raise ValueError("Khong co dong du lieu nao sau khi loc MakeBuy Code = B.")

    raw["_target_value"] = 0.0
    raw["_planned_sum"] = 0.0
    raw["_net_sum"] = 0.0
    raw["_offset_weeks"] = raw["Whse"].map(offset_map).fillna(0).astype(int)
    unknown_whse = sorted(set(raw.loc[~raw["Whse"].isin(offset_map.keys()), "Whse"].dropna().astype(str)))

    planned_etd_weeks = date_range_saturdays(first_etd_week, target_week)
    net_etd_weeks = date_range_saturdays(current_week, target_week)
    if current_week > target_week:
        raise ValueError("Target Week phai lon hon hoac bang Current Week.")

    # Process by offset instead of by row, much faster.
    for offset, idx in raw.groupby("_offset_weeks", sort=False).groups.items():
        target_src = target_week + timedelta(days=7 * int(offset))
        target_col = date_map.get(target_src)
        if target_col:
            raw.loc[idx, "_target_value"] = pd.to_numeric(raw.loc[idx, target_col], errors="coerce").fillna(0.0).values

        planned_cols = [date_map[d + timedelta(days=7 * int(offset))] for d in planned_etd_weeks if (d + timedelta(days=7 * int(offset))) in date_map]
        if planned_cols:
            raw.loc[idx, "_planned_sum"] = raw.loc[idx, planned_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0).sum(axis=1).values

        net_cols = [date_map[d + timedelta(days=7 * int(offset))] for d in net_etd_weeks if (d + timedelta(days=7 * int(offset))) in date_map]
        if net_cols:
            raw.loc[idx, "_net_sum"] = raw.loc[idx, net_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0).sum(axis=1).values

    key_cols = ["Item #", "Whse", "Coll. Class"]
    si_g = group_value(raw[raw["Data Type"] == "SHIPPABLE INV"], key_cols, "_target_value", "Base_SI")
    planned_g = group_value(raw[raw["Data Type"] == "PLANNED POS"], key_cols, "_planned_sum", "PlannedPO_Sum")
    firm_g = group_value(raw[raw["Data Type"] == "FIRM POS"], key_cols, "_target_value", "FirmPO_Target")
    net_g = group_value(raw[raw["Data Type"] == "NET FCST"], key_cols, "_net_sum", "NetFcst_Sum")
    ss_g = group_value(raw[raw["Data Type"] == "SAFETY STK"], key_cols, "_target_value", "SS_Wk3")

    base = raw[raw["Data Type"].isin(["SHIPPABLE INV", "PLANNED POS", "FIRM POS", "NET FCST", "SAFETY STK"])][key_cols].drop_duplicates()
    out = base.merge(si_g, on=key_cols, how="left").merge(planned_g, on=key_cols, how="left").merge(firm_g, on=key_cols, how="left").merge(net_g, on=key_cols, how="left").merge(ss_g, on=key_cols, how="left")
    for col in ["Base_SI", "PlannedPO_Sum", "FirmPO_Target", "NetFcst_Sum", "SS_Wk3"]:
        out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0.0)

    out["F Wk3"] = 0.0
    out["Sum of SI Wk3"] = out["Base_SI"] - out["PlannedPO_Sum"] - out["FirmPO_Target"]
    mask_335 = out["Whse"].astype(str) == "335"
    out.loc[mask_335, "Sum of SI Wk3"] = out.loc[mask_335, "Base_SI"] - out.loc[mask_335, "PlannedPO_Sum"] - out.loc[mask_335, "FirmPO_Target"] + out.loc[mask_335, "NetFcst_Sum"]
    out["Sum of SI-SS Wk3"] = out["Sum of SI Wk3"] - out["SS_Wk3"]
    out["Average of SS Wk3"] = out["SS_Wk3"]

    vendor_col = next((c for c in raw.columns if c.lower() == "vendor"), None)
    if vendor_col:
        vendor_map = raw.groupby(key_cols, dropna=False)[vendor_col].first().reset_index().rename(columns={vendor_col: "Vendor"})
        out = out.merge(vendor_map, on=key_cols, how="left")
    else:
        out["Vendor"] = ""

    out = out.rename(columns={"Item #": "Item", "Coll. Class": "ProdResourceID"})
    out["Item"] = out["Item"].map(normalize_item)
    out["Whse"] = out["Whse"].map(normalize_whse)

    f_wk3 = f_wk3.copy()
    f_wk3["Item"] = f_wk3["Item"].map(normalize_item)
    f_wk3["Whse"] = f_wk3["Whse"].map(normalize_whse)
    out = out.merge(f_wk3, on=["Item", "Whse"], how="left", suffixes=("", "_from_prod"))
    out["F Wk3"] = pd.to_numeric(out["F Wk3_from_prod"], errors="coerce").fillna(0.0)
    out = out.drop(columns=["F Wk3_from_prod"])
    output = out[OUTPUT_COLUMNS].drop_duplicates().copy()

    merge_debug = output.merge(f_wk3, on=["Item", "Whse"], how="left", indicator=True, suffixes=("", "_prod"))
    missing_f = merge_debug[merge_debug["_merge"] == "left_only"][["Item", "Whse", "ProdResourceID"]].drop_duplicates()

    build_debug = pd.DataFrame([
        ["Plan rows after MakeBuy B", len(raw)],
        ["Original first ETA week", fmt_date(first_original_week)],
        ["Original last ETA week", fmt_date(last_original_week)],
        ["First converted ETD week", fmt_date(first_etd_week)],
        ["TargetWeek", fmt_date(target_week)],
        ["CurrentWeek", fmt_date(current_week)],
        ["Planned POS ETD range", ", ".join(fmt_date(d) for d in planned_etd_weeks)],
        ["NET FCST ETD range", ", ".join(fmt_date(d) for d in net_etd_weeks)],
        ["F Wk3 source", "Production Schedule: only rows S/F/P = F, grouped by Item + Whse"],
        ["Rows output", str(len(output))],
        ["Rows without Production F match", str(len(missing_f))],
        ["F Wk3 total in optimizer input", str(float(output["F Wk3"].sum()))],
        ["Unknown Whse count", len(unknown_whse)],
        ["Unknown Whse list", ", ".join(unknown_whse)],
        ["Whse 335 SI logic", "SI(Target Week) - Planned POS(First ETD week -> Target Week) - Firm POS(Target Week) + Net Fcst(Current Week -> Target Week)"],
        ["Other Whse SI logic", "SI(Target Week) - Planned POS(First ETD week -> Target Week) - Firm POS(Target Week)"],
    ], columns=["Field", "Value"])

    offset_by_whse = raw[["Whse", "_offset_weeks"]].drop_duplicates().rename(columns={"_offset_weeks": "Used Offset Weeks"}).sort_values("Whse")
    return output, build_debug, missing_f, offset_by_whse


# ============================================================
# Step 4: Build optimizer input from converted PlanDetailTimeline
# ============================================================

def transform_converted_plan_to_optimizer_input(
    converted: pd.DataFrame,
    f_wk3: pd.DataFrame,
    target_week: date,
    current_week: date,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    raw = converted.copy()
    raw.columns = [str(c).strip() for c in raw.columns]

    required_base = ["Item #", "Whse", "Data Type", "Coll. Class", "MakeBuy Code"]
    missing_base = [c for c in required_base if c not in raw.columns]
    if missing_base:
        raise ValueError(f"Thieu cot bat buoc trong Plan converted: {', '.join(missing_base)}")

    raw["Data Type"] = clean_dtype(raw["Data Type"])
    raw["MakeBuy Code"] = raw["MakeBuy Code"].fillna("").astype(str).str.strip().str.upper()
    raw["Item #"] = raw["Item #"].map(normalize_item)
    raw["Whse"] = raw["Whse"].map(normalize_whse)
    raw["Coll. Class"] = raw["Coll. Class"].fillna("").astype(str).str.strip()

    vendor_col = next((c for c in raw.columns if c.lower() == "vendor"), None)

    raw = raw[raw["MakeBuy Code"] == "B"].copy()
    if raw.empty:
        raise ValueError("Khong co dong du lieu nao sau khi loc MakeBuy Code = B.")

    date_col_map = build_date_column_map(raw)
    target_col = date_col_map.get(target_week)
    if target_col is None:
        raise ValueError(f"Khong tim thay cot Target Week trong Plan converted: {fmt_date(target_week)}")

    all_week_dates = sorted(date_col_map.keys())
    first_week_date = min(all_week_dates)

    planned_weeks = date_range_saturdays(first_week_date, target_week)
    planned_missing = [fmt_date(d) for d in planned_weeks if d not in date_col_map]
    if planned_missing:
        raise ValueError("Thieu cot tuan cho Planned POS: " + ", ".join(planned_missing))

    if current_week > target_week:
        raise ValueError("Target Week phai lon hon hoac bang Current Week.")

    net_weeks = date_range_saturdays(current_week, target_week)
    net_missing = [fmt_date(d) for d in net_weeks if d not in date_col_map]
    if net_missing:
        raise ValueError("Thieu cot tuan cho NET FCST: " + ", ".join(net_missing))

    planned_cols = [date_col_map[d] for d in planned_weeks]
    net_cols = [date_col_map[d] for d in net_weeks]
    key_cols = ["Item #", "Whse", "Coll. Class"]

    si = raw[raw["Data Type"] == "SHIPPABLE INV"].copy()
    si["Base_SI"] = get_numeric(si, target_col)
    si_g = group_value(si, key_cols, "Base_SI", "Base_SI")

    planned = raw[raw["Data Type"] == "PLANNED POS"].copy()
    planned["PlannedPO_Sum"] = sum((get_numeric(planned, c) for c in planned_cols), start=pd.Series(0.0, index=planned.index))
    planned_g = group_value(planned, key_cols, "PlannedPO_Sum", "PlannedPO_Sum")

    firm = raw[raw["Data Type"] == "FIRM POS"].copy()
    firm["FirmPO_Target"] = get_numeric(firm, target_col)
    firm_g = group_value(firm, key_cols, "FirmPO_Target", "FirmPO_Target")

    net_fcst = raw[raw["Data Type"] == "NET FCST"].copy()
    net_fcst["NetFcst_Sum"] = sum((get_numeric(net_fcst, c) for c in net_cols), start=pd.Series(0.0, index=net_fcst.index))
    net_fcst_g = group_value(net_fcst, key_cols, "NetFcst_Sum", "NetFcst_Sum")

    ss = raw[raw["Data Type"] == "SAFETY STK"].copy()
    ss["SS_Wk3"] = get_numeric(ss, target_col)
    ss_g = group_value(ss, key_cols, "SS_Wk3", "SS_Wk3")

    base = raw[raw["Data Type"].isin(["SHIPPABLE INV", "PLANNED POS", "FIRM POS", "NET FCST", "SAFETY STK"])][key_cols].drop_duplicates()
    out = (
        base.merge(si_g, on=key_cols, how="left")
            .merge(planned_g, on=key_cols, how="left")
            .merge(firm_g, on=key_cols, how="left")
            .merge(net_fcst_g, on=key_cols, how="left")
            .merge(ss_g, on=key_cols, how="left")
    )

    for col in ["Base_SI", "PlannedPO_Sum", "FirmPO_Target", "NetFcst_Sum", "SS_Wk3"]:
        out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0.0)

    out["F Wk3"] = 0.0
    out["Sum of SI Wk3"] = out["Base_SI"] - out["PlannedPO_Sum"] - out["FirmPO_Target"]
    mask_335 = out["Whse"].astype(str) == "335"
    out.loc[mask_335, "Sum of SI Wk3"] = (
        out.loc[mask_335, "Base_SI"]
        - out.loc[mask_335, "PlannedPO_Sum"]
        - out.loc[mask_335, "FirmPO_Target"]
        + out.loc[mask_335, "NetFcst_Sum"]
    )
    out["Sum of SI-SS Wk3"] = out["Sum of SI Wk3"] - out["SS_Wk3"]
    out["Average of SS Wk3"] = out["SS_Wk3"]

    if vendor_col:
        vendor_map = raw.groupby(["Item #", "Whse", "Coll. Class"], dropna=False)[vendor_col].first().reset_index().rename(columns={vendor_col: "Vendor"})
        out = out.merge(vendor_map, on=key_cols, how="left")
    else:
        out["Vendor"] = ""

    out = out.rename(columns={"Item #": "Item", "Coll. Class": "ProdResourceID"})
    out["Item"] = out["Item"].map(normalize_item)
    out["Whse"] = out["Whse"].map(normalize_whse)

    f_wk3 = f_wk3.copy()
    f_wk3["Item"] = f_wk3["Item"].map(normalize_item)
    f_wk3["Whse"] = f_wk3["Whse"].map(normalize_whse)
    out = out.merge(f_wk3, on=["Item", "Whse"], how="left", suffixes=("", "_from_prod"))
    out["F Wk3"] = pd.to_numeric(out["F Wk3_from_prod"], errors="coerce").fillna(0.0)
    out = out.drop(columns=["F Wk3_from_prod"])

    output = out[OUTPUT_COLUMNS].drop_duplicates().copy()

    merge_debug = output.merge(f_wk3, on=["Item", "Whse"], how="left", indicator=True, suffixes=("", "_prod"))
    missing_f = merge_debug[merge_debug["_merge"] == "left_only"][["Item", "Whse", "ProdResourceID"]].drop_duplicates()

    debug_rows = [
        ["First week in converted file", fmt_date(first_week_date)],
        ["TargetWeek", fmt_date(target_week)],
        ["CurrentWeek", fmt_date(current_week)],
        ["Planned POS range", ", ".join(fmt_date(d) for d in planned_weeks)],
        ["NET FCST range", ", ".join(fmt_date(d) for d in net_weeks)],
        ["TargetWeek column found", str(target_col)],
        ["F Wk3 source", "Production Schedule: only rows S/F/P = F, grouped by Item + Whse"],
        ["Rows output", str(len(output))],
        ["Rows without Production F match", str(len(missing_f))],
        ["F Wk3 total in optimizer input", str(float(output["F Wk3"].sum()))],
        ["Whse 335 SI logic", "SI(Target Week) - Planned POS(First week -> Target Week) - Firm POS(Target Week) + Net Fcst(Current Week -> Target Week)"],
        ["Other Whse SI logic", "SI(Target Week) - Planned POS(First week -> Target Week) - Firm POS(Target Week)"],
    ]
    debug_df = pd.DataFrame(debug_rows, columns=["Field", "Value"])
    return output, debug_df, missing_f


# ============================================================
# Step 5: Optimizer logic from destination_change_optimizer_phase2only.py
# ============================================================

@dataclass
class PriorityRule:
    whse: str
    mode: str
    value: float


def normalize_pct(value) -> float:
    if value is None:
        return 0.0
    value = float(value)
    if value > 1 or value < -1:
        value = value / 100.0
    return value


def round_to_int_units(value: float) -> int:
    return int(round(float(value)))


def safe_ss_ratio(current_si: float, ss_target: float) -> float:
    if ss_target <= 0:
        if current_si > 0:
            return math.inf
        if current_si < 0:
            return -math.inf
        return 0.0
    return current_si / ss_target


def current_si_after(row: dict) -> int:
    return int(row["current_si"] + (row["final_f"] - row["orig_f"]))


def current_ss_after(row: dict) -> float:
    return safe_ss_ratio(current_si_after(row), row["ss_target"])


def compute_priority_target_final(row: dict, rule: PriorityRule) -> int:
    orig_f = row["orig_f"]
    current_si = row["current_si"]
    ss_target = row["ss_target"]
    if rule.mode == "SI":
        pct = max(0.0, min(1.0, float(rule.value)))
        target_si_after = current_si * (1.0 - pct)
        return max(0, round_to_int_units(orig_f + (target_si_after - current_si)))
    if rule.mode == "SS":
        target_si_after = ss_target * float(rule.value)
        return max(0, round_to_int_units(orig_f + (target_si_after - current_si)))
    return int(orig_f)


def build_rows(group: pd.DataFrame, item_rules: Dict[str, PriorityRule]) -> Tuple[List[dict], int]:
    rows = []
    for _, r in group.iterrows():
        row = {
            "item": r["Item"],
            "prod": r["ProdResourceID"],
            "whse": normalize_whse(r["Whse"]),
            "orig_f": round_to_int_units(r["F Wk3"]),
            "current_si": round_to_int_units(r["Current SI"]),
            "ss_target": float(r["Average of SS Wk3"]),
            "final_f": 0,
            "priority_rule_mode": "",
            "priority_rule_value": None,
            "priority_target_f_after": None,
        }
        rule = item_rules.get(row["whse"])
        if rule is not None:
            row["priority_rule_mode"] = rule.mode
            row["priority_rule_value"] = rule.value
            row["priority_target_f_after"] = compute_priority_target_final(row, rule)
        rows.append(row)
    return rows, round_to_int_units(group["Firm PO Total"].iloc[0])


def choose_priority_recipient(rows: List[dict], priority_indices: List[int]) -> Optional[int]:
    candidates = []
    for idx in priority_indices:
        row = rows[idx]
        target = row.get("priority_target_f_after")
        if target is None or row["final_f"] >= target:
            continue
        gap = target - row["final_f"]
        primary_metric = current_si_after(row) if row["priority_rule_mode"] == "SI" else current_ss_after(row)
        secondary_metric = current_ss_after(row) if row["priority_rule_mode"] == "SI" else current_si_after(row)
        candidates.append((gap, primary_metric, secondary_metric, row["whse"], idx))
    if not candidates:
        return None
    candidates.sort(key=lambda x: (-x[0], x[1], x[2], x[3]))
    return candidates[0][4]


def choose_lowest_ss_recipient(rows: List[dict], candidate_indices: List[int]) -> Optional[int]:
    candidates = []
    for idx in candidate_indices:
        after_si = current_si_after(rows[idx])
        ratio = safe_ss_ratio(after_si, rows[idx]["ss_target"])
        candidates.append((ratio, after_si, rows[idx]["whse"], idx))
    if not candidates:
        return None
    candidates.sort(key=lambda x: (x[0], x[1], x[2]))
    return candidates[0][3]


def allocate_item(group: pd.DataFrame, item_rules: Dict[str, PriorityRule]) -> pd.DataFrame:
    rows, total_f = build_rows(group, item_rules)
    remaining = total_f
    priority_indices = [i for i, r in enumerate(rows) if r["priority_rule_mode"]]
    non_priority_indices = [i for i, r in enumerate(rows) if not r["priority_rule_mode"]]

    while remaining > 0:
        idx = choose_priority_recipient(rows, priority_indices)
        if idx is None:
            break
        rows[idx]["final_f"] += 1
        remaining -= 1

    allocation_pool = non_priority_indices[:] if non_priority_indices else priority_indices[:]
    if remaining > 0 and allocation_pool:
        import heapq
        heap = []
        for idx in allocation_pool:
            after_si = current_si_after(rows[idx])
            ratio = safe_ss_ratio(after_si, rows[idx]["ss_target"])
            heapq.heappush(heap, (ratio, after_si, rows[idx]["whse"], idx))
        while remaining > 0 and heap:
            _, _, _, idx = heapq.heappop(heap)
            rows[idx]["final_f"] += 1
            remaining -= 1
            after_si = current_si_after(rows[idx])
            ratio = safe_ss_ratio(after_si, rows[idx]["ss_target"])
            heapq.heappush(heap, (ratio, after_si, rows[idx]["whse"], idx))

    out = group.copy().reset_index(drop=True)
    out["F Wk3 Original"] = out["F Wk3"].round().astype(int)
    out["F Wk3 After Destination Change"] = [r["final_f"] for r in rows]
    out["Net Destination Change"] = out["F Wk3 After Destination Change"] - out["F Wk3 Original"]
    out["Current SI After"] = out["Current SI"] + out["Net Destination Change"]
    out["SS % After"] = out.apply(lambda r: safe_ss_ratio(float(r["Current SI After"]), float(r["Average of SS Wk3"])), axis=1)
    out["Remaining Unallocated PO"] = total_f - int(out["F Wk3 After Destination Change"].sum())
    out["Priority Rule Mode"] = [r["priority_rule_mode"] for r in rows]
    out["Priority Rule Value"] = [r["priority_rule_value"] for r in rows]
    out["Priority Target F After"] = [r["priority_target_f_after"] for r in rows]

    if int(out["F Wk3 Original"].sum()) != int(out["F Wk3 After Destination Change"].sum()):
        raise ValueError(f"Item {out['Item'].iloc[0]}: tong firm khong bao toan.")
    return out


def prepare_optimizer_input(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    required = ["Item", "ProdResourceID", "Whse", "F Wk3", "Sum of SI Wk3", "Average of SS Wk3"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Optimizer input thieu cot: {missing}")
    if not [c for c in df.columns if "vendor" in c.lower()]:
        df[VENDOR_FALLBACK_COL] = ""

    df["Item"] = df["Item"].map(normalize_item)
    df = df[df["Item"] != ""].copy()
    df["Whse"] = df["Whse"].map(normalize_whse)
    for c in ["F Wk3", "Sum of SI Wk3", "Average of SS Wk3"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    if "Sum of SI-SS Wk3" in df.columns:
        df["Sum of SI-SS Wk3"] = pd.to_numeric(df["Sum of SI-SS Wk3"], errors="coerce").fillna(0)
    else:
        df["Sum of SI-SS Wk3"] = pd.NA

    df["Current SI"] = df["Sum of SI Wk3"]
    df["Current SS%"] = df.apply(lambda r: safe_ss_ratio(float(r["Current SI"]), float(r["Average of SS Wk3"])), axis=1)
    totals = df.groupby("Item", as_index=False)["F Wk3"].sum().rename(columns={"F Wk3": "Firm PO Total"})
    df = df.merge(totals, on="Item", how="left")
    return df


def build_detail_output(detail: pd.DataFrame) -> pd.DataFrame:
    preferred = [
        "Item", "ProdResourceID", "Whse", "F Wk3", "Sum of SI Wk3", "Sum of SI-SS Wk3", "Average of SS Wk3",
        "Current SI", "Current SS%", "Firm PO Total", "F Wk3 Original", "F Wk3 After Destination Change",
        "Net Destination Change", "Current SI After", "SS % After", "Remaining Unallocated PO",
        "Priority Rule Mode", "Priority Rule Value", "Priority Target F After",
    ]
    vendor_cols = [c for c in detail.columns if "vendor" in c.lower()]
    final_cols = [c for c in preferred if c in detail.columns]
    for c in detail.columns:
        if c not in final_cols and c not in vendor_cols:
            final_cols.append(c)
    for c in vendor_cols:
        if c not in final_cols:
            final_cols.append(c)
    return detail[final_cols].copy()


def build_summary(detail: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, g in detail.groupby("Item", sort=True):
        rows.append({
            "Item": g["Item"].iloc[0],
            "ProdResourceID": g["ProdResourceID"].iloc[0],
            "Firm PO Total": int(g["Firm PO Total"].iloc[0]),
            "Total F Before": int(g["F Wk3 Original"].sum()),
            "Total F After": int(g["F Wk3 After Destination Change"].sum()),
            "Min SI After": int(g["Current SI After"].min()),
            "Max SI After": int(g["Current SI After"].max()),
            "Min SS % After": float(g["SS % After"].min()),
            "Max SS % After": float(g["SS % After"].max()),
        })
    return pd.DataFrame(rows)


def apply_zero_ss_fallback_equalization(detail_full: pd.DataFrame) -> pd.DataFrame:
    """Fallback equalization pass for items with multiple warehouses where SS target is zero.

    Business intent:
    - Keep the main allocation logic unchanged.
    - If an item has 2+ warehouses with Average of SS Wk3 = 0,
      rebalance only that zero-SS group using SI After instead of SS%.
    - Preserve the total F Wk3 After Destination Change for the item.

    This pass is intentionally conservative:
    - It never touches warehouses with SS target > 0.
    - It only redistributes within the zero-SS subset.
    """
    if detail_full is None or detail_full.empty:
        return detail_full

    detail = detail_full.copy()
    if "Average of SS Wk3" not in detail.columns:
        return detail

    for item, idx in detail.groupby("Item", sort=False).groups.items():
        idx = list(idx)
        ss = pd.to_numeric(detail.loc[idx, "Average of SS Wk3"], errors="coerce").fillna(0.0).to_numpy()
        zero_mask = ss <= 0
        zero_idx = [idx[i] for i, is_zero in enumerate(zero_mask) if is_zero]
        if len(zero_idx) < 2:
            continue

        current_si = pd.to_numeric(detail.loc[zero_idx, "Current SI"], errors="coerce").fillna(0.0).to_numpy(dtype=float)
        orig_f = pd.to_numeric(detail.loc[zero_idx, "F Wk3 Original"], errors="coerce").fillna(0.0).to_numpy(dtype=float)
        final_f = pd.to_numeric(detail.loc[zero_idx, "F Wk3 After Destination Change"], errors="coerce").fillna(0.0).to_numpy(dtype=float)

        total_final_f = int(round(final_f.sum()))
        if total_final_f <= 0:
            continue

        # Equalize by SI After within the zero-SS subset.
        # Target common SI After = average of the subset after preserving its total firm quantity.
        target_si_after = (current_si.sum() + (total_final_f - orig_f.sum())) / len(zero_idx)
        desired_final_f = orig_f + (target_si_after - current_si)

        rounded = _sum_preserving_round(
            desired_final_f,
            total_final_f,
            lower=np.zeros(len(zero_idx), dtype=int),
            upper=np.full(len(zero_idx), max(total_final_f, 1), dtype=int),
        ).astype(int)

        detail.loc[zero_idx, "F Wk3 After Destination Change"] = rounded
        detail.loc[zero_idx, "Net Destination Change"] = detail.loc[zero_idx, "F Wk3 After Destination Change"] - detail.loc[zero_idx, "F Wk3 Original"]
        detail.loc[zero_idx, "Current SI After"] = detail.loc[zero_idx, "Current SI"] + detail.loc[zero_idx, "Net Destination Change"]
        detail.loc[zero_idx, "SS % After"] = detail.loc[zero_idx].apply(
            lambda r: safe_ss_ratio(float(r["Current SI After"]), float(r["Average of SS Wk3"])), axis=1
        )

    return detail


def run_optimizer(optimizer_input: pd.DataFrame, priority_rules: Optional[Dict[str, PriorityRule]] = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
    priority_rules = priority_rules or {}
    df = prepare_optimizer_input(optimizer_input)
    records = []
    for _, group in df.groupby("Item", sort=True):
        item_whse = set(group["Whse"].astype(str).tolist())
        item_rules = {wh: rule for wh, rule in priority_rules.items() if wh in item_whse}
        allocated = allocate_item(group.copy(), item_rules)
        records.extend(allocated.to_dict("records"))
    detail_full = pd.DataFrame.from_records(records) if records else pd.DataFrame()
    if not detail_full.empty and int(detail_full["F Wk3 Original"].sum()) != int(detail_full["F Wk3 After Destination Change"].sum()):
        raise ValueError("Sai tong Firm PO toan file.")

    # Fallback equalization pass: if an item has multiple warehouses with SS=0,
    # rebalance only that subset using SI After so the zero-SS warehouses do not
    # become heavily imbalanced when coverage is disabled.
    detail_full = apply_zero_ss_fallback_equalization(detail_full)

    return build_detail_output(detail_full), build_summary(detail_full)


# ============================================================
# Output writer
# ============================================================

def autofit(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def style_sheet(ws):
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    autofit(ws)


def write_excel_output(
    output_path: str,
    detail: pd.DataFrame,
    summary: pd.DataFrame,
    optimizer_input: pd.DataFrame,
    debug_sheets: Dict[str, pd.DataFrame],
):
    output_path = ensure_unique_output_path(output_path)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        detail.to_excel(writer, sheet_name="Optimized Data", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        optimizer_input.to_excel(writer, sheet_name="Optimizer Input Debug", index=False)
        for name, df in debug_sheets.items():
            safe_name = name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
        logic = pd.DataFrame([
            ["Input PlanDetailTimeline", "Raw CSV, code tu dong bo qua metadata lines truoc dong Item #"],
            ["ETA -> ETD", "Plan week value is shifted earlier by Used Offset Weeks from DueDateCalc/legacy mapping"],
            ["F Wk3", "Production Schedule, only S/F/P = F, grouped by Item + Whse at Target Week"],
            ["Current SI", "= Sum of SI Wk3"],
            ["Current SS%", "= Current SI / Average of SS Wk3"],
            ["F Wk3 After", "Firm cuoi cung cua kho sau destination change"],
            ["Net Destination Change", "= F Wk3 After - F Wk3 Original"],
            ["Current SI After", "= Current SI + Net Destination Change"],
            ["SS % After", "= Current SI After / Average of SS Wk3"],
            ["Priority mode SI", "Target la cover theo % huong toi SI = 0"],
            ["Priority mode SS", "Target %SS dung cong thuc SI / SS"],
            ["Allocation rule", "Sau priority, luon chon kho co SS% After thap nhat"],
        ], columns=["Field", "Meaning"])
        logic.to_excel(writer, sheet_name="Logic", index=False)

        wb = writer.book
        for ws in wb.worksheets:
            style_sheet(ws)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    header = ws.cell(1, cell.column).value
                    if header in ["Current SS%", "SS % After", "Priority Rule Value"] and isinstance(cell.value, (int, float)) and not math.isinf(cell.value):
                        cell.number_format = "0.0%"
    return output_path


# ============================================================
# One-shot process
# ============================================================

def process_files(
    plan_detail_csv: str,
    production_schedule_csv: str,
    due_date_calc_xlsx: str,
    output_path: str,
    target_week: date,
    current_week: Optional[date] = None,
    priority_rules: Optional[Dict[str, PriorityRule]] = None,
    offset_mode: str = "legacy_compatible",
) -> str:
    if current_week is None:
        current_week = saturday_of_current_week()

    offset_map, offset_debug = load_due_date_offsets(due_date_calc_xlsx, offset_mode=offset_mode)
    f_wk3, prod_debug = load_fwk3_from_production(production_schedule_csv, target_week)
    optimizer_input, build_debug, missing_f, plan_offset_debug = build_optimizer_input_direct_from_plan(
        plan_detail_csv, offset_map, f_wk3, target_week, current_week
    )
    convert_debug = pd.DataFrame([
        ["Conversion mode", "Direct fast path, equivalent to ETA -> ETD shift without writing converted workbook"],
        ["Materialized converted workbook", "No"],
    ], columns=["Field", "Value"])
    detail, summary = run_optimizer(optimizer_input, priority_rules=priority_rules)

    run_debug = pd.DataFrame([
        ["PlanDetailTimeline", plan_detail_csv],
        ["Production Schedule", production_schedule_csv],
        ["DueDateCalc", due_date_calc_xlsx],
        ["TargetWeek", fmt_date(target_week)],
        ["CurrentWeek", fmt_date(current_week)],
        ["Offset mode", offset_mode],
        ["Output detail rows", len(detail)],
        ["Total F original", int(detail["F Wk3 Original"].sum()) if not detail.empty else 0],
        ["Total F after", int(detail["F Wk3 After Destination Change"].sum()) if not detail.empty else 0],
    ], columns=["Field", "Value"])

    debug_sheets = {
        "Run Debug": run_debug,
        "DueDate Offset Debug": offset_debug,
        "Plan Convert Debug": convert_debug,
        "Plan Offset Debug": plan_offset_debug,
        "Production Debug": prod_debug,
        "Build Debug": build_debug,
        "Missing F Match": missing_f,
    }
    return write_excel_output(output_path, detail, summary, optimizer_input, debug_sheets)


# ============================================================
# Tkinter UI
# ============================================================

def ask_priority_rules(warehouses: List[str]) -> Dict[str, PriorityRule]:
    rules: Dict[str, PriorityRule] = {}
    if simpledialog is None or tk is None:
        return rules
    root = tk.Tk()
    root.withdraw()
    valid = set(warehouses)
    raw = simpledialog.askstring(
        "Kho uu tien",
        "Nhap mot hoac nhieu kho uu tien, cach nhau bang dau phay.\n"
        "Vi du: 11,17,28\n"
        "De trong neu khong co kho uu tien.",
    )
    if not raw:
        root.destroy()
        return rules
    selected = []
    for x in raw.split(","):
        wh = normalize_whse(x)
        if wh and wh in valid and wh not in selected:
            selected.append(wh)
    for wh in selected:
        mode = simpledialog.askstring(
            f"Rule cho kho {wh}",
            "Nhap mode cho kho nay:\n- SI: cover mot phan toi SI = 0\n- SS: cover toi mot muc %SS nhat dinh",
        )
        if not mode:
            continue
        mode = mode.strip().upper()
        if mode == "SI":
            val = simpledialog.askfloat(f"% cover SI cho kho {wh}", "Vi du 50 = cover 50%, 100 = full toi SI=0")
            if val is not None:
                rules[wh] = PriorityRule(whse=wh, mode="SI", value=normalize_pct(val))
        elif mode == "SS":
            val = simpledialog.askfloat(f"Target %SS cho kho {wh}", "Vi du 0 = SI bang SS; -100 = SI bang 0; 100 = SI bang 2 lan SS")
            if val is not None:
                rules[wh] = PriorityRule(whse=wh, mode="SS", value=normalize_pct(val))
    root.destroy()
    return rules


class UnifiedApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Destination Change Unified Flow")
        self.root.geometry("980x680")
        self.root.minsize(920, 620)
        self.plan_var = tk.StringVar()
        self.prod_var = tk.StringVar()
        self.due_var = tk.StringVar()
        self.output_var = tk.StringVar()
        default_current = saturday_of_current_week()
        self.current_var = tk.StringVar(value=fmt_date(default_current))
        self.target_var = tk.StringVar(value=fmt_date(default_current + timedelta(days=14)))
        self.offset_mode_var = tk.StringVar(value="legacy_compatible")
        self.status_var = tk.StringVar(value="San sang")
        self.build_ui()

    def build_ui(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("Title.TLabel", font=("Segoe UI", 15, "bold"))
        style.configure("Big.TButton", font=("Segoe UI", 12, "bold"), padding=(16, 12))
        outer = ttk.Frame(self.root, padding=16)
        outer.pack(fill="both", expand=True)
        ttk.Label(outer, text="Destination Change Unified Flow", style="Title.TLabel").pack(anchor="w", pady=(0, 12))

        files = ttk.LabelFrame(outer, text="Input / Output files")
        files.pack(fill="x", pady=(0, 12))
        self.file_row(files, 0, "PlanDetailTimeline raw CSV:", self.plan_var, self.browse_plan)
        self.file_row(files, 2, "Production Schedule raw CSV:", self.prod_var, self.browse_prod)
        self.file_row(files, 4, "DueDateCalc Excel:", self.due_var, self.browse_due)
        self.file_row(files, 6, "Output Excel:", self.output_var, self.browse_output, save=True)
        files.columnconfigure(1, weight=1)

        dates = ttk.LabelFrame(outer, text="Week setup")
        dates.pack(fill="x", pady=(0, 12))
        ttk.Label(dates, text="Target Week / Wk3:").grid(row=0, column=0, sticky="w", padx=10, pady=10)
        ttk.Entry(dates, textvariable=self.target_var, width=20).grid(row=0, column=1, sticky="w", padx=10, pady=10)
        ttk.Label(dates, text="Current Week:").grid(row=1, column=0, sticky="w", padx=10, pady=10)
        ttk.Entry(dates, textvariable=self.current_var, width=20).grid(row=1, column=1, sticky="w", padx=10, pady=10)
        ttk.Label(dates, text="ETA -> ETD offset mode:").grid(row=2, column=0, sticky="w", padx=10, pady=10)
        ttk.Combobox(dates, textvariable=self.offset_mode_var, values=["legacy_compatible", "due_date"], state="readonly", width=24).grid(row=2, column=1, sticky="w", padx=10, pady=10)

        info = ttk.LabelFrame(outer, text="Logic summary")
        info.pack(fill="x", pady=(0, 12))
        logic_text = (
            "1) PlanDetailTimeline raw CSV -> auto skip metadata -> ETA converted to ETD by warehouse offset.\n"
            "2) Production Schedule -> F Wk3 from S/F/P = F only, target week column must match user input.\n"
            "3) Build optimizer input with SI, SI-SS, SS, Vendor + F Wk3.\n"
            "4) Run destination change optimizer and write final Excel with debug sheets."
        )
        ttk.Label(info, text=logic_text, justify="left", wraplength=900).pack(anchor="w", padx=10, pady=10)

        actions = ttk.Frame(outer)
        actions.pack(fill="x")
        ttk.Label(actions, textvariable=self.status_var).pack(side="left")
        ttk.Button(actions, text="RUN", command=self.run, style="Big.TButton").pack(side="right")

    def file_row(self, parent, row, label, var, command, save=False):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=10, pady=(8, 2))
        ttk.Entry(parent, textvariable=var).grid(row=row + 1, column=0, columnspan=2, sticky="ew", padx=10, pady=(0, 8))
        ttk.Button(parent, text="Save As" if save else "Browse", command=command).grid(row=row + 1, column=2, sticky="e", padx=10, pady=(0, 8))

    def browse_plan(self):
        p = filedialog.askopenfilename(title="Select PlanDetailTimeline CSV", filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if p:
            self.plan_var.set(p)
            self.default_output_from_plan(p)

    def browse_prod(self):
        p = filedialog.askopenfilename(title="Select Production Schedule CSV", filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if p:
            self.prod_var.set(p)

    def browse_due(self):
        p = filedialog.askopenfilename(title="Select DueDateCalc Excel", filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")])
        if p:
            self.due_var.set(p)

    def browse_output(self):
        p = filedialog.asksaveasfilename(title="Save output", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if p:
            self.output_var.set(p)

    def default_output_from_plan(self, plan_path):
        if not self.output_var.get().strip():
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_var.set(str(Path(plan_path).with_name(f"destination_change_unified_{stamp}.xlsx")))

    def run(self):
        try:
            plan = self.plan_var.get().strip()
            prod = self.prod_var.get().strip()
            due = self.due_var.get().strip()
            out = self.output_var.get().strip()
            if not plan or not prod or not due or not out:
                raise ValueError("Ban can chon day du 4 file/path: PlanDetailTimeline, Production Schedule, DueDateCalc, Output.")
            target_week = parse_user_date(self.target_var.get())
            current_week = parse_user_date(self.current_var.get())
            self.status_var.set("Dang doc warehouse list de nhap priority...")
            self.root.update_idletasks()
            # Build a preliminary F list to know warehouses. If this fails, main run will show error.
            prod_f, _ = load_fwk3_from_production(prod, target_week)
            whses = sorted(prod_f["Whse"].dropna().astype(str).unique().tolist(), key=lambda x: (len(x), x))
            rules = ask_priority_rules(whses)
            self.status_var.set("Dang xu ly full flow...")
            self.root.update_idletasks()
            final_path = process_files(
                plan_detail_csv=plan,
                production_schedule_csv=prod,
                due_date_calc_xlsx=due,
                output_path=out,
                target_week=target_week,
                current_week=current_week,
                priority_rules=rules,
                offset_mode=self.offset_mode_var.get(),
            )
            self.status_var.set("Hoan tat")
            messagebox.showinfo("Done", f"Da tao file output:\n{final_path}")
        except Exception as e:
            self.status_var.set("Co loi")
            messagebox.showerror("Error", str(e))


# ============================================================
# CLI
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Destination Change Unified Flow")
    parser.add_argument("--plan", help="PlanDetailTimeline raw CSV")
    parser.add_argument("--production", help="Production Schedule raw CSV")
    parser.add_argument("--due", help="DueDateCalc Excel")
    parser.add_argument("--target-week", help="Target week, e.g. 5/23/2026")
    parser.add_argument("--current-week", help="Current week, e.g. 5/9/2026")
    parser.add_argument("--output", help="Output Excel path")
    parser.add_argument("--offset-mode", choices=["legacy_compatible", "due_date"], default="legacy_compatible")
    parser.add_argument("--no-gui", action="store_true", help="Run CLI only")
    args = parser.parse_args()

    if args.no_gui or all([args.plan, args.production, args.due, args.target_week, args.output]):
        if not all([args.plan, args.production, args.due, args.target_week, args.output]):
            raise SystemExit("CLI mode requires --plan --production --due --target-week --output")
        final = process_files(
            plan_detail_csv=args.plan,
            production_schedule_csv=args.production,
            due_date_calc_xlsx=args.due,
            output_path=args.output,
            target_week=parse_user_date(args.target_week),
            current_week=parse_user_date(args.current_week) if args.current_week else None,
            priority_rules={},
            offset_mode=args.offset_mode,
        )
        print(f"Done: {final}")
        return

    if tk is None:
        raise SystemExit("Tkinter khong kha dung. Hay chay CLI voi --no-gui va cac tham so input.")
    root = tk.Tk()
    UnifiedApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
